from tkinter import *
from tkinter import messagebox, filedialog
from tkinter import Menu
from tkinter import ttk
from tkinter.ttk import Combobox
import openpyxl
import pandas as pd
import modul

ablak = Tk()
ablak.title('Könyvek nyilvántartása')
ablak.geometry('700x400+100+100')
ablak.resizable(False, False)
ablak.configure(bg='#66B2FF')

def uj_oldal():
    uj_ablak = Toplevel(ablak)
    uj_ablak.title('Adatok')
    uj_ablak.geometry('1100x420+100+100')
    uj_ablak.configure(bg='#66B2FF')
    def megnyit():
        filename = filedialog.askopenfilename(title="Fájl megnyitása", filetype=(('xlsx files','*.xlsx'),('All files', '*.*')))
        if filename:
            try:
                filename = r'{}'.format(filename)
                df = pd.read_excel(filename)
            except:
                messagebox.showerror('Hiba', "Nem elérhető a fájl!")

            tree.delete(*tree.get_children())

            tree['column'] = list(df.columns)
            tree['show'] = 'headings'

            for oszlop in tree['column']:
                tree.heading(oszlop, text=oszlop)

            df_rows = df.to_numpy().tolist()
            for sor in df_rows:
                tree.insert('', 'end', values=sor)

    frame = Frame(uj_ablak)
    frame.pack(pady=25)

    tree = ttk.Treeview(frame)
    tree.pack()

    megnyit_gomb = Button(uj_ablak, text='Megnyit', bg='#fff', width=15, height=1, command=megnyit)
    megnyit_gomb.pack(padx=10, pady=20)

    uj_ablak.mainloop()

#Menü
menubar = Menu(ablak)
ablak.config(menu=menubar)

file_menu = Menu(menubar, tearoff=0)
menubar.add_cascade(label="File",menu=file_menu)

file_menu.add_command(label='Új oldal', command=uj_oldal)
file_menu.add_separator()
file_menu.add_command(label='Kilépés', command=ablak.destroy)

#Modul meghívása
modul.fajl()
def kuld():
    sznev = sz_nev.get()
    kcim = k_cim.get()
    khossz = k_hossz.get()
    knyelv = k_nyelv.get()
    rido = r_ido.get()
    ertekeless = ertekeles.get()
    leirass = leiras.get(1.0, 'end')

    fajl = openpyxl.load_workbook('Adatok.xlsx')
    sheet = fajl.active
    sheet.cell(column=1, row=sheet.max_row+1, value=sznev)
    sheet.cell(column=2, row=sheet.max_row, value=kcim)
    sheet.cell(column=3, row=sheet.max_row, value=khossz)
    sheet.cell(column=4, row=sheet.max_row, value=knyelv)
    sheet.cell(column=5, row=sheet.max_row, value=rido)
    sheet.cell(column=6, row=sheet.max_row, value=ertekeless)
    sheet.cell(column=7, row=sheet.max_row, value=leirass)
    fajl.save(r'Adatok.xlsx')

    messagebox.showinfo('info', 'Hozzáadva!')

    sz_nev.delete(first=0, last=100)
    k_cim.delete(first=0, last=100)
    k_hossz.delete(first=0, last=100)
    r_ido.delete(first=0, last=100)
    k_nyelv.delete(first=0, last=100)
    ertekeles.delete(first=0, last=100)
    leiras.delete(1.0, 'end')

def torol():
    sz_nev.delete(first=0, last=100)
    k_cim.delete(first=0, last=100)
    k_hossz.delete(first=0, last=100)
    r_ido.delete(first=0, last=100)
    k_nyelv.delete(first=0, last=100)
    ertekeles.delete(first=0, last=100)
    leiras.delete(1.0, 'end')

#Főoldal
Frame(ablak, width=600, height=300, bg='#CCE5FF').place(x=50, y=50)
Label(ablak, text='Szerző neve', font='calibri 12 bold', bg='#CCE5FF').place(x=60, y=70)
Label(ablak, text='Könyv címe', font='calibri 12 bold', bg='#CCE5FF').place(x=60, y=110)
Label(ablak, text='Könyv hossza', font='calibri 12 bold', bg='#CCE5FF').place(x=60, y=150)
Label(ablak, text='Könyv nyelve', font='calibri 12 bold', bg='#CCE5FF').place(x=60, y=190)
Label(ablak, text='Ráfordított idő', font='calibri 12 bold', bg='#CCE5FF').place(x=60, y=230)
Label(ablak, text='Rövid leírás', font='calibri 12 bold', bg='#CCE5FF').place(x=400, y=70)
Label(ablak, text='Értékelés', font='calibri 12 bold', bg='#CCE5FF').place(x=60, y=270)

szerzonev = StringVar()
konyvcim = StringVar()
konyvhossz = StringVar()
leiras = StringVar()
raf_ido = StringVar()
leiras = StringVar()

sz_nev = Entry(ablak,textvariable=szerzonev, width=30, bd=2)
sz_nev.pack()
sz_nev.place(x=170, y=75)
k_cim = Entry(ablak,textvariable=konyvcim, width=30, bd=2)
k_cim.pack()
k_cim.place(x=170, y=115)
k_hossz = Entry(ablak,textvariable=konyvhossz, width=30, bd=2)
k_hossz.pack()
k_hossz.place(x=170, y=155)
r_ido = Entry(ablak,textvariable=raf_ido, width=30, bd=2)
r_ido.pack
r_ido.place(x=170, y=235)

k_nyelv = Combobox(ablak, values=['Magyar', 'Angol'], width=14)
k_nyelv.pack()
k_nyelv.place(x=170, y=195)
ertekeles = Combobox(ablak, values=['1','2','3','4','5'], width=14)
ertekeles.pack()
ertekeles.place(x=170, y=270)

leiras = Text(ablak, width=32, height=15, bd=2, font='calibri 10')
leiras.pack()
leiras.place(x=400, y=100)

Button(ablak, text='Küld', bg='#fff', width=15, height=1, command=kuld).place(x=50, y=360)
Button(ablak, text='Töröl', bg='#fff', width=15, height=1, command=torol).place(x=170, y=360)

ablak.mainloop()